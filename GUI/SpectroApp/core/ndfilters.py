"""
core/ndfilters.py — Neutral-density (ND) filter correction.

Different shots may have been recorded with different ND filters in front of
the side-SRS spectrometer (column "side-SRS ND" of Final.xlsx: the optical
density OD of the filter stack). To compare absolute signal levels between
shots, spectra must be corrected for the filter transmission.

Two correction levels, from most to least accurate:
  1. MEASURED transmission curve: the user supplies, for each OD value used
     during the campaign, the manufacturer's Excel file giving transmission
     (%) versus wavelength (nm). The curve is linearly interpolated onto the
     calibration wavelength axis and the spectrum is divided by T(lambda).
  2. THEORETICAL fallback: if no file is given for an OD value, the flat
     Beer-Lambert value T = 10^(-OD) is used (with a clear warning in the UI).

File format (as provided by filter manufacturers, e.g. Thorlabs NEx series):
anywhere in the first sheet, a header cell "Wavelength (nm)" with
"Transmission (%)" in the column immediately to its right, followed by
numeric rows. The parser scans for this pair so leading metadata columns or
rows are ignored.

The registry {od_value_str: file_path} is persisted in config.json via the
Session; parsed curves are cached in memory only.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np

# In-memory cache: file path (str) -> (wl_nm ndarray, T_fraction ndarray)
_CURVE_CACHE: dict[str, tuple[np.ndarray, np.ndarray]] = {}


def parse_nd_file(path) -> tuple[np.ndarray, np.ndarray]:
    """Parse an ND transmission Excel file.

    Returns (wavelength_nm, transmission_fraction), sorted by wavelength.
    Raises ValueError with an explicit message if the format is not
    recognised.
    """
    import openpyxl
    path = Path(path)
    if not path.exists():
        raise ValueError(f"file not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def _norm(v):
        return str(v).strip().lower() if v is not None else ""

    # Locate the "wavelength"/"transmission" header pair
    wl_col = tr_col = start_row = None
    for r, row in enumerate(rows):
        for c, v in enumerate(row):
            h = _norm(v)
            if h.startswith("wavelength"):
                # transmission expected in a nearby column of the same row
                for c2 in range(c + 1, min(c + 3, len(row))):
                    if _norm(row[c2]).startswith("transmission"):
                        wl_col, tr_col, start_row = c, c2, r + 1
                        break
            if wl_col is not None:
                break
        if wl_col is not None:
            break
    if wl_col is None:
        raise ValueError(
            "no 'Wavelength (nm)' / 'Transmission (%)' header pair found "
            "in the first sheet")

    pct = "%" in _norm(rows[start_row - 1][tr_col])
    wl, tr = [], []
    for row in rows[start_row:]:
        if len(row) <= max(wl_col, tr_col):
            continue
        w, t = row[wl_col], row[tr_col]
        if w is None or t is None:
            continue
        try:
            wl.append(float(w))
            tr.append(float(t))
        except (TypeError, ValueError):
            continue
    if len(wl) < 2:
        raise ValueError("fewer than 2 numeric (wavelength, transmission) "
                         "rows found")
    wl = np.asarray(wl, float)
    tr = np.asarray(tr, float)
    if pct or np.nanmax(tr) > 1.5:   # values look like percentages
        tr = tr / 100.0
    order = np.argsort(wl)
    return wl[order], tr[order]


def get_curve(path) -> tuple[np.ndarray, np.ndarray]:
    """Cached access to a parsed transmission curve."""
    key = str(path)
    if key not in _CURVE_CACHE:
        _CURVE_CACHE[key] = parse_nd_file(path)
    return _CURVE_CACHE[key]


def clear_cache():
    _CURVE_CACHE.clear()


def transmission_on_axis(od_value, wl_axis, nd_files: dict) -> tuple[
        np.ndarray, str]:
    """Transmission T(lambda) of the filter stack for a given OD value,
    evaluated on the calibration wavelength axis.

    Parameters
    ----------
    od_value : the OD value from the shotbook (float-able). 0/None -> no
        attenuation.
    wl_axis  : (n_px,) wavelength axis in nm.
    nd_files : {od_value_str: file_path} registry.

    Returns
    -------
    (T, source) with T an (n_px,) array of transmission fractions and
    source one of 'none', 'file', 'theory:10^-OD', or 'file+clamp' when the
    axis extends beyond the measured curve (edge values are held constant
    there).
    """
    wl_axis = np.asarray(wl_axis, float)
    try:
        od = float(od_value)
    except (TypeError, ValueError):
        od = 0.0
    if od == 0.0:
        return np.ones_like(wl_axis), "none"

    key = _od_key(od)
    path = (nd_files or {}).get(key)
    if path and Path(path).exists():
        try:
            wl, tr = get_curve(path)
            T = np.interp(wl_axis, wl, tr)   # clamps outside the range
            T = np.clip(T, 1e-12, None)      # never divide by zero
            clamped = (wl_axis.min() < wl.min() - 1e-9
                       or wl_axis.max() > wl.max() + 1e-9)
            return T, ("file+clamp" if clamped else "file")
        except ValueError:
            pass  # unreadable file -> theoretical fallback
    return np.full_like(wl_axis, 10.0 ** (-od)), "theory:10^-OD"


def correction_factor(od_value, wl_axis, nd_files: dict) -> tuple[
        np.ndarray, str]:
    """Multiplicative correction 1/T(lambda) to recover the pre-filter
    signal. Returns (factor, source) — see transmission_on_axis."""
    T, source = transmission_on_axis(od_value, wl_axis, nd_files)
    return 1.0 / T, source


def _od_key(od: float) -> str:
    """Canonical dict key for an OD value ('1', '2', '0.5', ...)."""
    return f"{od:g}"


def od_values_in_metadata(metadata: dict) -> list[str]:
    """Sorted list of distinct non-zero OD keys found in the shotbook
    metadata ({shot: {'nd': value, ...}})."""
    vals = set()
    for m in (metadata or {}).values():
        nd = m.get("nd")
        try:
            od = float(nd)
        except (TypeError, ValueError):
            continue
        if od != 0.0:
            vals.add(_od_key(od))
    return sorted(vals, key=float)


def od_in_file(path) -> float | None:
    """Try to read the filter's optical density from the file's own header
    text, e.g. a cell containing 'ARC: 350-700 nm, OD: 1.0'. Manufacturers
    (Thorlabs NEx, etc.) write the OD in a metadata cell near the top of the
    sheet. Returns the OD as a float, or None if not found.

    Scans only the first ~40 rows so a stray 'OD' in the numeric data cannot
    be mistaken for the header value.
    """
    import re
    import openpyxl
    path = Path(path)
    if not path.exists():
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None
    ws = wb[wb.sheetnames[0]]
    pat = re.compile(r"\bOD\s*[:=]?\s*([0-9]+(?:\.[0-9]+)?)", re.I)
    od = None
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r >= 40:
            break
        for v in row:
            if v is None:
                continue
            m = pat.search(str(v))
            if m:
                od = float(m.group(1))
                break
        if od is not None:
            break
    wb.close()
    return od


def scan_nd_folder(folder, wanted_od_keys=None) -> tuple[dict, list]:
    """Scan a folder for ND transmission Excel files and auto-associate each
    to an OD value.

    For every .xlsx/.xls file that both (a) parses as a transmission curve
    and (b) exposes an OD in its header text, the file is mapped to the
    corresponding OD key. When two files claim the same OD, the first one
    (alphabetical) wins and the clash is reported.

    Parameters
    ----------
    folder : directory to scan.
    wanted_od_keys : optional iterable of OD keys actually used in the
        shotbook (e.g. ['1', '2', '3']); files whose OD is not in this set
        are reported as 'unused' rather than mapped.

    Returns
    -------
    (nd_files, reports) where nd_files is {od_key: path} and reports is a
    list of (level, message) with level in {'ok', 'warn', 'info'}.
    """
    folder = Path(folder)
    reports = []
    if not folder.is_dir():
        return {}, [("warn", f"ND folder not found: {folder}")]
    wanted = set(wanted_od_keys) if wanted_od_keys is not None else None
    nd_files: dict[str, str] = {}
    candidates = sorted([p for p in folder.iterdir()
                         if p.suffix.lower() in (".xlsx", ".xls")
                         and not p.name.startswith("~$")])
    for p in candidates:
        od = od_in_file(p)
        if od is None:
            continue  # not an ND datasheet (no OD in header) — silently skip
        # confirm it is really a transmission curve
        try:
            parse_nd_file(p)
        except ValueError:
            reports.append(("info", f"'{p.name}': OD {od:g} found but no "
                                    f"transmission table — skipped."))
            continue
        key = _od_key(od)
        if wanted is not None and key not in wanted:
            reports.append(("info", f"'{p.name}': OD {od:g} not used in the "
                                    f"shotbook — ignored."))
            continue
        if key in nd_files:
            reports.append(("warn", f"OD {key}: several files match "
                                    f"('{Path(nd_files[key]).name}' kept, "
                                    f"'{p.name}' ignored)."))
            continue
        nd_files[key] = str(p)
        reports.append(("ok", f"OD {key} <- '{p.name}'"))
    if not nd_files:
        reports.append(("warn", "No ND datasheet recognised in this folder "
                                "(files must contain an 'OD: x.x' header and "
                                "a Wavelength/Transmission table)."))
    return nd_files, reports
