"""
core/metadata.py — Metadonnees de campagne depuis Final.xlsx.

Lit pour chaque shot : profil de pulse, cible, energie 2w, configuration
fibres (colonne 'side-SRS fibrePos'). Sert a :
  - construire AUTOMATIQUEMENT les groupes (cible x profil), avec filtre
    optionnel en energie (centre +/- tolerance %) — remplace le classement
    manuel des shots du notebook, en restant verifiable (tableau d'assignation
    affiche avant creation) ;
  - resoudre la configuration angulaire d'un shot depuis la colonne dediee
    (generalisable aux campagnes futures), avec repli sur les plages codees
    en dur du pipeline.

Les colonnes sont trouvees par leur EN-TETE (ligne 1), avec repli sur les
lettres de colonnes de la campagne actuelle (C, D, F, G, BH). L'energie reste
lue par sf.load_energy_table (chemin de code d'origine) ; la valeur E2w lue
ici ne sert qu'a l'affichage et au filtrage des groupes — un ecart entre les
deux serait un bug de lecture, pas une difference scientifique.
"""
from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import openpyxl

# En-tetes attendus (insensible casse/espaces) -> repli lettre de colonne.
# Le repli lettre ne sert que si l'en-tete est introuvable ET n'est utilise
# que pour la campagne d'origine ; pour l'ND (colonne tres variable d'une
# campagne a l'autre) il n'y a PAS de repli lettre : on detecte, ou l'on
# laisse l'utilisateur choisir.
_COLS = {
    "shot": (("shot n", "shot"), "C"),
    "profile": (("pulse profile",), "D"),
    "e2w": (("2w e (j)", "2w e"), "F"),
    "target": (("target",), "G"),
    "fiberpos": (("side-srs fibrepos", "side-srs fiberpos"), "BH"),
}


def _norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _find_header_row(ws, max_scan=6):
    """Return the 1-based index of the header row (the first row, within the
    first few, that contains a 'shot' column). Falls back to row 1."""
    for r in range(1, max_scan + 1):
        vals = [_norm(c.value) for c in ws[r] if c.value is not None]
        if any(v.startswith("shot") for v in vals):
            return r
    return 1


def nd_columns(excel_path) -> list[tuple[str, str]]:
    """Every column of the shotbook whose header names an ND (optical
    density) value, as (header_text, column_letter), in sheet order.

    Different diagnostics have their own ND (SRS ND, side-SRS ND,
    Resolved-SSRS ND, Back SOP ND, ...); which one drives the multi-fiber
    side-scattering spectrometer is campaign-dependent, so the choice is left
    to the user (with a sensible default from :func:`auto_nd_column`)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = _find_header_row(ws)
    out = []
    for c in ws[hr]:
        if c.value is None:
            continue
        h = _norm(c.value)
        # 'nd' as a standalone word (avoids matching 'grand', 'and', ...)
        if re.search(r"\bnd\b", h):
            out.append((re.sub(r"\s+", " ", str(c.value)).strip(),
                        c.column_letter))
    wb.close()
    return out


def auto_nd_column(nd_cols: list[tuple[str, str]]) -> str | None:
    """Best default ND column among the shotbook's ND columns. The multi-fiber
    side-scattering spectrometer is the (S)SRS one, so prefer, in order: the
    exact legacy 'side-SRS ND'; a side-scattering SRS column ('SSRS'); a
    side/SRS column that is not another diagnostic (SOP/SBS/GOI); else the
    first ND column. This is only a guess — the user picks the real one."""
    if not nd_cols:
        return None
    norm = [(h, _norm(h)) for h, _ in nd_cols]
    for h, n in norm:
        if n == "side-srs nd":
            return h
    for h, n in norm:
        if "ssrs" in n:
            return h
    for h, n in norm:
        if "srs" in n and "sop" not in n and "sbs" not in n and "goi" not in n:
            return h
    for h, n in norm:
        if "side" in n and "sop" not in n:
            return h
    return nd_cols[0][0]


def load_final_table(excel_path, nd_column: str | None = None) -> dict:
    """{shot: {'profile', 'profile_raw', 'target', 'target_raw', 'si_pct',
               'e2w', 'fiberpos', 'nd'}}

    ``nd_column`` selects which shotbook column holds the ND value (its
    header text). None = auto-detect (:func:`auto_nd_column`)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = _find_header_row(ws)
    # localisation des colonnes par en-tete (ligne d'en-tete detectee)
    header = {c.column_letter: _norm(c.value) for c in ws[hr]
              if c.value is not None}
    letters = {}
    for key, (names, fallback) in _COLS.items():
        found = None
        for letter, h in header.items():
            if any(h == n or h.startswith(n) for n in names):
                found = letter
                break
        letters[key] = found or fallback

    # ND column: explicit choice > auto-detect > none
    nd_letter = None
    nd_cols = [(re.sub(r"\s+", " ", str(c.value)).strip(), c.column_letter)
               for c in ws[hr]
               if c.value is not None and re.search(r"\bnd\b", _norm(c.value))]
    target_hdr = nd_column or auto_nd_column(nd_cols)
    if target_hdr is not None:
        tnorm = _norm(target_hdr)
        for h, letter in nd_cols:
            if _norm(h) == tnorm:
                nd_letter = letter
                break
    if nd_letter:
        letters["nd"] = nd_letter

    idx = {k: openpyxl.utils.column_index_from_string(v) - 1
           for k, v in letters.items()}
    has_nd = "nd" in idx

    table = {}
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        need = max(idx.values())
        if len(row) <= need:
            row = tuple(row) + (None,) * (need + 1 - len(row))
        try:
            shot = int(row[idx["shot"]])
        except (TypeError, ValueError):
            continue
        prof_raw = row[idx["profile"]]
        targ_raw = row[idx["target"]]
        e2w = row[idx["e2w"]]
        fpos = row[idx["fiberpos"]]
        nd_raw = row[idx["nd"]] if has_nd else None
        try:
            e2w = float(e2w) if e2w is not None else None
        except (TypeError, ValueError):
            e2w = None
        try:
            nd_val = float(nd_raw) if nd_raw is not None else None
        except (TypeError, ValueError):
            nd_val = None   # e.g. 'x' means the diagnostic was not used
        table[shot] = {
            "nd": nd_val,
            "profile_raw": str(prof_raw) if prof_raw is not None else "",
            "profile": normalize_profile(prof_raw),
            "target_raw": str(targ_raw) if targ_raw is not None else "",
            "target": normalize_target(targ_raw),
            "si_pct": parse_si_pct(targ_raw),
            "e2w": e2w,
            "fiberpos": str(fpos).strip() if fpos is not None else None,
        }
    wb.close()
    return table


def normalize_profile(raw) -> str:
    """'10 90'->'10/90', '30 70'->'30/70', '2ns FT'->'2ns' ; le reste est
    conserve tel quel (nettoye)."""
    if raw is None:
        return "?"
    s = _norm(raw)
    m = re.fullmatch(r"(\d{1,2})\s*[/ ]\s*(\d{1,2})", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    if s.startswith("2ns"):
        return "2ns"
    return re.sub(r"\s+", " ", str(raw)).strip()


def normalize_target(raw) -> str:
    """Nettoie la cible (espaces multiples/finaux, '+5 Si' -> '+ 5 Si') pour
    que '25 CH + 25 Kapton ' et '25 CH + 25 Kapton' forment UN groupe."""
    if raw is None:
        return "?"
    s = re.sub(r"\s+", " ", str(raw)).strip()
    s = re.sub(r"\+\s*(\d+)\s*Si", r"+ \1 Si", s, flags=re.IGNORECASE)
    return s


def parse_si_pct(raw):
    """% Si depuis la cible : '+ N Si' -> N ; cible CH/Kapton/PP sans Si -> 0 ;
    sinon None (ex : 'Al frame')."""
    if raw is None:
        return None
    s = str(raw)
    m = re.search(r"\+\s*(\d+)\s*Si", s, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    if re.search(r"\bCH\b|Kapton|PP\b", s, flags=re.IGNORECASE):
        return 0
    return None


# ── Groupes automatiques ─────────────────────────────────────────────────────
PALETTE = ["#e6194B", "#3cb44b", "#4363d8", "#f58231", "#911eb4", "#008080",
           "#9A6324", "#800000", "#808000", "#000075", "#f032e6", "#2f4f4f"]


def build_auto_groups(table: dict, available_shots: set[str] | None = None,
                      e_center: float | None = None, tol_pct: float = 10.0,
                      min_size: int = 2, only_available: bool = True):
    """Construit les groupes (cible normalisee x profil normalise).

    Retourne (groups, assign_rows, excluded_rows) ou :
      groups   : liste au format campaign.py (name, color, si_pct, profile,
                 shots, plus 'target'),
      assign_rows  : (shot, cible, profil, E2w, config, nom_groupe),
      excluded_rows: (shot, cible, profil, E2w, raison).
    Aucune invention : chaque exclusion est motivee et affichable."""
    buckets: dict[tuple, list] = {}
    assign, excluded = [], []
    for shot, m in sorted(table.items()):
        key = f"shot{shot:03d}"
        reason = None
        if m["target"] in ("?", "") or m["profile"] in ("?", ""):
            reason = "target or profile missing from Final.xlsx"
        elif e_center is not None:
            if m["e2w"] is None:
                reason = "E2w missing"
            elif abs(m["e2w"] - e_center) > e_center * tol_pct / 100.0:
                reason = (f"E2w = {m['e2w']:.1f} J outside the window "
                          f"{e_center:.0f} J ± {tol_pct:.0f} %")
        if reason is None and only_available and available_shots is not None \
                and key not in available_shots:
            reason = "image missing from the folder"
        if reason is not None:
            excluded.append((shot, m["target"], m["profile"], m["e2w"], reason))
            continue
        buckets.setdefault((m["target"], m["profile"]), []).append((shot, m))

    groups = []
    small = 0
    for i, ((target, profile), members) in enumerate(sorted(buckets.items())):
        if len(members) < min_size:
            for shot, m in members:
                excluded.append((shot, target, profile, m["e2w"],
                                 f"group '{target} — {profile}' too small "
                                 f"({len(members)} < {min_size})"))
            small += 1
            continue
        name = f"{target} — {profile}"
        if e_center is not None:
            name += f" @ {e_center:.0f}J±{tol_pct:.0f}%"
        shots = [f"shot{s:03d}" for s, _ in members]
        groups.append({
            "name": name,
            "color": PALETTE[len(groups) % len(PALETTE)],
            "si_pct": members[0][1]["si_pct"],
            "profile": profile,
            "target": target,
            "shots": shots,
        })
        for shot, m in members:
            assign.append((shot, target, profile, m["e2w"],
                           m["fiberpos"], name))
    return groups, assign, excluded
