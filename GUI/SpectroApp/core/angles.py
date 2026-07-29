"""
core/angles.py — Configurations angulaires depuis les fichiers Excel.

Automatise ce qui etait fait a la main : associer (phi, theta) a chaque fibre.

Sources :
  1. Fichier "structure" (ex : 251102_EstimateStructureCoordinates.xlsx),
     feuille 1 : table Lettre->phi (col A=phi, B=lettre Alpha, C=lettre Delta)
     et tables Port->theta (cols E/F pour Delta, H/I pour Alpha).
  2. Fichiers de position (ex : SidescatterFibrePos_Config3_d.xlsx) :
     col A = Channel 1..80, col D = Quadrant, col E = Arm (lettre),
     col F = Port (numero).

Regle de signe : theta -> -theta pour le quadrant choisi par NEGATE_QUADRANT.
VALIDATION FAITE : avec negate='alpha', les trois configurations codees en dur
du pipeline (config3_d/e/f) sont reproduites EXACTEMENT (240 fibres, delta=0).
NB : la description orale de la regle mentionnait Delta ; les valeurs validees
du pipeline correspondent a Alpha. Le parametre est expose pour rester
verifiable.

Les configurations construites ici sont indexees par fibre PHYSIQUE
(Channel n = fibre physique n-1), exactement comme sf.FIBER_CONFIGS ; le pont
indice-detecteur -> fibre physique reste fait par l'inversion phys = 79 - i,
identique a sf.get_fiber_angles.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import numpy as np
import openpyxl

from core import spectro_functions as sf

NEGATE_QUADRANT = "alpha"   # regle validee contre le pipeline (voir docstring)

_CFG_FILE_RE = re.compile(r"SidescatterFibrePos_(.+)\.xlsx$", re.IGNORECASE)


def load_structure_tables(path) -> dict:
    """Lit les 4 tables de correspondance du fichier structure."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    phi_alpha, phi_delta, th_alpha, th_delta = {}, {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 9:
            row = tuple(row) + (None,) * (9 - len(row))
        phi, l_a, l_d = row[0], row[1], row[2]
        if phi is not None and l_a:
            phi_alpha[str(l_a).strip().upper()] = float(phi)
        if phi is not None and l_d:
            phi_delta[str(l_d).strip().upper()] = float(phi)
        if row[4] is not None and row[5] is not None:
            th_delta[int(row[5])] = float(row[4])
        if row[7] is not None and row[8] is not None:
            th_alpha[int(row[8])] = float(row[7])
    if not (phi_alpha and phi_delta and th_alpha and th_delta):
        raise ValueError(
            "Tables de structure incompletes : attendu col A=phi, B=lettre "
            "Alpha, C=lettre Delta, E/F=theta/port Delta, H/I=theta/port Alpha.")
    return {"phi_alpha": phi_alpha, "phi_delta": phi_delta,
            "th_alpha": th_alpha, "th_delta": th_delta}


def _theta_lookup(table: dict, port: int):
    """theta pour un port, avec extrapolation UNIQUEMENT si la table est une
    progression arithmetique stricte (cas observe : pas de -5.74 deg/port).
    Retourne (theta | None, extrapole: bool).
    Justification : la table du fichier structure s'arrete au port 12 alors
    que certaines configurations utilisent des ports au-dela ; la valeur
    extrapolee au port 14 (9.64 deg) reproduit exactement la valeur validee
    du pipeline pour config3_e."""
    if port in table:
        return table[port], False
    ports = sorted(table)
    if len(ports) < 3:
        return None, False
    steps = [(table[ports[i + 1]] - table[ports[i]])
             / (ports[i + 1] - ports[i]) for i in range(len(ports) - 1)]
    step = steps[0]
    if any(abs(s - step) > 1e-6 for s in steps):
        return None, False           # table non arithmetique : pas d'invention
    p0 = ports[0]
    return table[p0] + step * (port - p0), True


def build_config_from_files(structure: dict, positions_path,
                            negate=NEGATE_QUADRANT):
    """Retourne (fibres_dict, rapport) ou fibres_dict = {fibre_physique:
    (phi, theta)} au format exact de sf.FIBER_CONFIGS[...]['fibres']."""
    wb = openpyxl.load_workbook(positions_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    fibres, xyz, quadrants = {}, {}, {}
    issues = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            ch = int(row[0])
        except (TypeError, ValueError):
            continue
        if not (1 <= ch <= 80):
            continue
        quad = str(row[3]).strip().lower() if row[3] else ""
        arm = str(row[4]).strip().upper() if row[4] else ""
        port = row[5]
        if not quad or not arm or port is None:
            issues.append(f"Channel {ch}: incomplete quadrant/arm/port — "
                          f"fiber left without angles (NaN).")
            continue
        try:
            port = int(port)
        except (TypeError, ValueError):
            issues.append(f"Channel {ch}: unreadable port ({port!r}).")
            continue
        if quad.startswith("alpha"):
            phi = structure["phi_alpha"].get(arm)
            th, extrap = _theta_lookup(structure["th_alpha"], port)
        elif quad.startswith("delta"):
            phi = structure["phi_delta"].get(arm)
            th, extrap = _theta_lookup(structure["th_delta"], port)
        else:
            issues.append(f"Channel {ch}: unknown quadrant '{row[3]}'.")
            continue
        if phi is None:
            issues.append(f"Channel {ch}: arm letter '{arm}' absent from the "
                          f"phi table ({quad}).")
            continue
        if th is None:
            issues.append(f"Channel {ch}: port {port} absent from the theta "
                          f"table ({quad}) and table not extrapolable.")
            continue
        if extrap:
            issues.append(f"Channel {ch}: port {port} beyond the theta "
                          f"table — extrapolated with the constant step "
                          f"({th:+.2f} deg before sign).")
        # Angles BRUTS : la regle de signe est appliquee a l'affichage, comme
        # pour les plans de campagne. C'est ce qui manquait ici et qui laissait
        # le menu des quadrants vide sur les campagnes lues par ce chemin.
        quad_norm = "alpha" if quad.startswith("alpha") else "delta"
        quadrants[ch - 1] = quad_norm
        fibres[ch - 1] = (float(phi), float(th))   # Channel n -> physique n-1
        xyz[ch - 1] = tuple(float(c) for c in
                            xyz_from_angles(phi, th))
    found = sorted(set(quadrants.values()))
    rule = [q for q in found if q.startswith(str(negate or ""))] \
        if negate else []
    ent = {"fibres": fibres, "xyz": xyz, "numbering": "physical",
           "source": Path(positions_path).name, "quadrant": quadrants,
           "sign_rules": normalise_sign_rules({"theta": rule, "phi": []})}
    return ent, issues


def compare_to_builtin(name: str, fibres: dict):
    """Si la config existe en dur dans le pipeline, compare fibre a fibre.
    Retourne (statut, message)."""
    key = name.lower()
    if key not in sf.FIBER_CONFIGS:
        return "nouveau", f"'{name}': new configuration (absent from the pipeline)."
    fibres = entry(fibres)["fibres"] if isinstance(fibres, dict) \
        and "fibres" in fibres else fibres
    ref = sf.FIBER_CONFIGS[key]["fibres"]
    if set(ref) != set(fibres):
        only_ref = sorted(set(ref) - set(fibres))
        only_new = sorted(set(fibres) - set(ref))
        return "ecart", (f"'{name}': fibers present differ from the pipeline "
                         f"(pipeline only: {only_ref[:6]}, Excel only: "
                         f"{only_new[:6]}).")
    dmax = max(max(abs(ref[k][0] - fibres[k][0]),
                   abs(ref[k][1] - fibres[k][1])) for k in ref)
    if dmax < 1e-9:
        return "identique", (f"'{name}': identical to the validated pipeline "
                             f"configuration (80 fibers, max gap = 0).")
    bad = [k for k in ref if abs(ref[k][0] - fibres[k][0]) > 1e-9
           or abs(ref[k][1] - fibres[k][1]) > 1e-9]
    return "ecart", (f"'{name}': DISCREPANCY with the pipeline on {len(bad)} "
                     f"fibers (e.g. physical {bad[:5]}, max gap {dmax:.4g}).")


# ── Registre (persiste dans workspace/angles.json) ───────────────────────────
def registry_path(workspace) -> Path:
    return Path(workspace) / "angles.json"


def load_registry(workspace) -> dict:
    """Registre des configurations angulaires.

    Deux formes d'entree cohabitent (compatibilite ascendante) :
      * ancienne : {nom: {fibre: (phi, theta)}}
      * nouvelle : {nom: {"fibres": {...}, "xyz": {...},
                          "numbering": "physical"|"extraction",
                          "source": nom_de_fichier}}
    Les accesseurs `entry`, `fibres_of`, `xyz_of` masquent la difference.
    """
    p = registry_path(workspace)
    if not p.exists():
        return {}
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}
    out = {}
    for name, d in raw.items():
        try:
            fib = {int(k): tuple(v) for k, v in d["fibres"].items()}
            xyz = {int(k): tuple(v) for k, v in (d.get("xyz") or {}).items()}
            quad = {int(k): v
                    for k, v in (d.get("quadrant") or {}).items()}
            out[name] = {"fibres": fib, "xyz": xyz,
                         "numbering": d.get("numbering", "physical"),
                         "source": d.get("source", ""),
                         "quadrant": quad,
                         "sign_rules": d.get("sign_rules")}
        except Exception:
            continue
    return out


def save_registry(workspace, registry: dict):
    raw = {}
    for name, e in registry.items():
        ent = entry(e)
        raw[name] = {
            "fibres": {str(k): list(v) for k, v in ent["fibres"].items()},
            "xyz": {str(k): list(v) for k, v in (ent["xyz"] or {}).items()},
            "numbering": ent["numbering"],
            "source": ent.get("source", ""),
            "quadrant": {str(k): v
                         for k, v in (ent.get("quadrant") or {}).items()},
            "sign_rules": ent.get("sign_rules") or {},
        }
    registry_path(workspace).write_text(
        json.dumps(raw, indent=1), encoding="utf-8")


# ── Accesseurs tolerants aux deux formes d'entree ────────────────────────────
def entry(e) -> dict:
    """Normalise une entree de registre."""
    if isinstance(e, dict) and "fibres" in e:
        return {"fibres": e.get("fibres") or {}, "xyz": e.get("xyz") or {},
                "numbering": e.get("numbering", "physical"),
                "source": e.get("source", ""),
                "quadrant": e.get("quadrant") or {},
                "sign_rules": normalise_sign_rules(e.get("sign_rules"))}
    return {"fibres": e or {}, "xyz": {}, "numbering": "physical",
            "source": "", "quadrant": {},
            "sign_rules": normalise_sign_rules(None)}


# ── Regle de signe, choisie par l'utilisateur ────────────────────────────────
# Les angles sont desormais stockes BRUTS (tels que lus dans le fichier de
# reference) et le changement de signe est applique a l'affichage. Une regle
# figee dans le code ne pouvait pas suivre : elle etait validee sur le
# quadrant alpha d'une campagne, alors que la convention depend du montage.
# On expose donc le choix — quel angle (theta ou phi), pour quel(s)
# quadrant(s) parmi ceux REELLEMENT presents dans le fichier.

DEFAULT_SIGN_RULES = {"theta": [NEGATE_QUADRANT], "phi": []}


def normalise_sign_rules(rules) -> dict:
    """{'theta': [quadrants], 'phi': [quadrants]} — tolerant aux entrees
    partielles, aux None et aux chaines seules."""
    if rules is None:
        return {k: list(v) for k, v in DEFAULT_SIGN_RULES.items()}
    out = {}
    for k in ("theta", "phi"):
        v = (rules or {}).get(k) or []
        if isinstance(v, str):
            v = [v]
        out[k] = sorted({str(q).strip().lower() for q in v if str(q).strip()})
    return out


# LA conversion angles -> cartesien est celle de votre code d'origine,
# `sf._fiber_angles_to_xyz`, et il n'y en a pas d'autre. Je l'avais reecrite
# ici en v25 en croyant reproduire la convention spherique usuelle
# (x=sinθcosφ, y=sinθsinφ, z=cosθ) : c'est faux. La votre echange les roles
# de phi et theta, traite le quadrant negatif par un azimut miroir, et place
# l'axe polaire sur y — de sorte que l'angle a la retrodiffusion vaut
# exactement theta. Comme `build_config_from_map` stockait des coordonnees
# calculees avec ma formule, `fig_sphere` les utilisait au lieu d'appeler la
# votre : c'est l'origine de tout ce que nous avons chasse depuis.

def xyz_from_angles(phi_deg, theta_deg, axis=None):
    """(x, y, z) — delegue a la conversion validee du pipeline.

    `axis` n'est garde que pour la compatibilite des appels : il n'y a qu'une
    convention et elle ne se choisit pas.
    """
    return sf._fiber_angles_to_xyz(np.asarray(phi_deg, float),
                                   np.asarray(theta_deg, float))


def quadrants_of(registry: dict, name: str) -> list[str]:
    """Quadrants effectivement rencontres dans une configuration importee.
    Sert a peupler les menus : on ne propose que ce qui existe."""
    resolved = _resolve_name(name, registry)
    if resolved is None:
        return []
    q = entry(registry[resolved]).get("quadrant") or {}
    return sorted({str(v).strip().lower() for v in q.values() if v})


def _apply_sign(ent: dict, sign_rules=None):
    """(fibres, xyz) apres application de la regle de signe.

    x, y, z sont TOUJOURS recalcules depuis les angles signes : garder les
    coordonnees d'origine tout en inversant un angle donnerait une sphere
    incoherente avec sa propre carte 2D.
    """
    ent = entry(ent)
    rules = normalise_sign_rules(
        sign_rules if sign_rules is not None else ent.get("sign_rules"))
    quad = ent.get("quadrant") or {}
    fibres, xyz = {}, {}
    for k, (phi, th) in ent["fibres"].items():
        q = str(quad.get(k, quad.get(str(k), ""))).strip().lower()
        if q and q in rules["theta"]:
            th = -th
        if q and q in rules["phi"]:
            phi = -phi
        fibres[k] = (float(phi), float(th))
        xyz[k] = tuple(float(c) for c in xyz_from_angles(phi, th))
    return fibres, xyz


def fibres_of(registry: dict, name: str) -> dict:
    return entry(registry.get(name, {}))["fibres"]


def xyz_of(registry: dict, name: str) -> dict:
    return entry(registry.get(name, {}))["xyz"]


def scan_position_files(folder) -> dict:
    """{nom_config: chemin} pour tous les SidescatterFibrePos_*.xlsx."""
    out = {}
    d = Path(folder)
    if not d.is_dir():
        return out
    for f in sorted(d.iterdir()):
        m = _CFG_FILE_RE.match(f.name)
        if m:
            out[m.group(1)] = f
    return out


# ── Acces unifie (pipeline en priorite, puis registre) ───────────────────────
def known_configs(registry: dict) -> list[str]:
    builtin = list(sf.FIBER_CONFIGS.keys())
    custom = [c for c in registry if c.lower() not in sf.FIBER_CONFIGS]
    return builtin + sorted(custom)


def _resolve_name(config_name: str, registry: dict) -> str | None:
    if config_name in registry:
        return config_name
    for k in registry:
        if k.lower() == str(config_name).lower():
            return k
    return None


def get_fiber_angles_any(config_name: str, registry: dict, n_fibers=None,
                         sign_rules=None, numbering=None):
    """Angles indexes par indice d'extraction. Pour les configs du pipeline,
    appelle sf.get_fiber_angles (chemin exact). Pour les configs du registre,
    applique la MEME inversion physique."""
    key = config_name.lower()
    if key in sf.FIBER_CONFIGS:
        return sf.get_fiber_angles(key)
    resolved = _resolve_name(config_name, registry)
    if resolved is None:
        raise ValueError(f"Unknown configuration '{config_name}' "
                         f"(neither pipeline nor imported files).")
    config_name = resolved
    ent = entry(registry[config_name])
    cfg, _ = _apply_sign(ent, sign_rules)
    numb = numbering or ent["numbering"]
    if n_fibers is None:
        n_fibers = sf.N_FIBERS
    phis = np.full(n_fibers, np.nan)
    thetas = np.full(n_fibers, np.nan)
    for det_i in range(n_fibers):
        key = _index_for(det_i, n_fibers, numb)
        if key in cfg:
            phis[det_i] = cfg[key][0]
            thetas[det_i] = cfg[key][1]
    return phis, thetas


def _index_for(det_i: int, n_fibers: int, numbering: str):
    """Cle de la table pour un indice d'extraction.

    'physical'   : la table est indexee par fibre physique — on applique
                   l'inversion du faisceau (identique au pipeline).
    'extraction' : la table est deja dans l'ordre d'extraction — aucune
                   inversion (cas des fichiers ou la numerotation suit
                   directement les traces du detecteur).
    """
    if str(numbering).lower().startswith("extract"):
        return int(det_i)
    return int(sf.physical_fiber_index(det_i, n_fibers))


def get_fiber_xyz_any(config_name: str, registry: dict, n_fibers=None,
                      sign_rules=None, numbering=None):
    """Coordonnees (3, n_fibres) indexees par indice d'extraction, ou None si
    la configuration n'en fournit pas. La regle de signe est appliquee ici :
    les coordonnees suivent donc toujours les angles affiches."""
    name = _resolve_name(config_name, registry)
    if name is None:
        return None
    ent = entry(registry[name])
    _, xyz = _apply_sign(ent, sign_rules)
    if not xyz:
        return None
    numb = numbering or ent["numbering"]
    if n_fibers is None:
        n_fibers = sf.N_FIBERS
    out = np.full((3, n_fibers), np.nan)
    for det_i in range(n_fibers):
        key = _index_for(det_i, n_fibers, numb)
        if key in xyz:
            out[:, det_i] = xyz[key]
    return out


# =============================================================================
# Format « table plate » — un seul fichier, une ligne par fibre
# =============================================================================
# Campagnes recentes : plus de fichier structure + fichiers de position, mais
# un unique classeur donnant directement, pour chaque fibre, ses angles et sa
# position. Exemple d'en-tetes (fiber_config_Fauvel_b.xlsx) :
#
#   Fiber # | Group | Config type | Arm | Port | Phi (deg) | Theta (deg) |
#   Role | x | y | z
#
# Le lecteur ci-dessous reconnait ces colonnes par leur nom (insensible a la
# casse, aux accents et a la ponctuation), et laisse l'utilisateur corriger la
# correspondance depuis l'interface si une campagne future change les intitules.

FLAT_FIELDS = ["fiber", "phi", "theta", "x", "y", "z",
               "group", "arm", "port", "role"]

FLAT_ALIASES = {
    "fiber": ["fiber", "fiber #", "fiber number", "fiber id", "fibre",
              "fibre #", "channel", "chan", "num", "n", "id", "index"],
    "phi":   ["phi", "phi deg", "phi degrees", "φ", "phi_deg", "azimuth"],
    "theta": ["theta", "theta deg", "theta degrees", "θ", "theta_deg",
              "polar", "elevation"],
    "x": ["x", "x pos", "pos x", "x (mm)", "xcoord"],
    "y": ["y", "y pos", "pos y", "y (mm)", "ycoord"],
    "z": ["z", "z pos", "pos z", "z (mm)", "zcoord"],
    "group": ["group", "groupe", "quadrant"],
    "arm":   ["arm", "bras", "letter", "lettre"],
    "port":  ["port", "position"],
    "role":  ["role"],
}


def _norm_header(h) -> str:
    """Normalise un intitule de colonne pour la comparaison."""
    if h is None:
        return ""
    t = str(h).strip().lower()
    for a, b in (("é", "e"), ("è", "e"), ("ê", "e"), ("°", ""), ("º", "")):
        t = t.replace(a, b)
    t = re.sub(r"[\(\)\[\]\.,;:_/\\-]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _guess_columns(headers: list[str]) -> dict:
    """{champ: intitule} pour les colonnes reconnues automatiquement."""
    norm = {h: _norm_header(h) for h in headers}
    guess = {}
    used = set()
    for field in FLAT_FIELDS:
        aliases = FLAT_ALIASES[field]
        # 1) egalite stricte, alias les plus longs d'abord
        for alias in sorted(aliases, key=len, reverse=True):
            for h in headers:
                if h in used:
                    continue
                if norm[h] == alias:
                    guess[field] = h
                    used.add(h)
                    break
            if field in guess:
                break
        if field in guess:
            continue
        # 2) l'intitule commence par l'alias ("phi deg" -> "phi")
        for alias in sorted(aliases, key=len, reverse=True):
            for h in headers:
                if h in used:
                    continue
                if norm[h].startswith(alias + " ") or norm[h] == alias:
                    guess[field] = h
                    used.add(h)
                    break
            if field in guess:
                break
    return guess


def sniff_flat_table(path, sheet=None) -> dict:
    """Inspecte un classeur : feuilles, en-tetes, correspondance devinee.
    Ne fait aucune hypothese silencieuse : tout est renvoye pour affichage."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = list(wb.sheetnames)
    name = sheet if sheet in sheets else sheets[0]
    ws = wb[name]
    header_row, headers = None, []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12,
                                         values_only=True), start=1):
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        if len(cells) < 2:
            continue
        cand = [str(c).strip() for c in row]
        g = _guess_columns([c for c in cand if c and c != "None"])
        if "fiber" in g and ("phi" in g or "x" in g):
            header_row, headers = i, [c if c and c != "None" else ""
                                      for c in cand]
            break
    if header_row is None:                    # repli : premiere ligne remplie
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12,
                                             values_only=True), start=1):
            cells = [c for c in row if c is not None]
            if len(cells) >= 2:
                header_row = i
                headers = [str(c).strip() if c is not None else ""
                           for c in row]
                break
    n_rows = sum(1 for _ in ws.iter_rows(min_row=(header_row or 1) + 1))
    wb.close()
    heads = [h for h in headers if h]
    return {"sheets": sheets, "sheet": name, "header_row": header_row or 1,
            "headers": heads, "guess": _guess_columns(heads),
            "n_rows": n_rows}


def load_flat_table(path, sheet=None, col_map=None, numbering="physical",
                    theta_sign=1.0, phi_sign=1.0, derive_xyz=True,
                    n_fibers=None):
    """Lit une table plate et construit une entree de registre.

    Parametres
    ----------
    col_map    : {champ: intitule de colonne} — remplace la detection
                 automatique pour les champs fournis.
    numbering  : 'physical'   -> la colonne fibre est un numero de fibre
                                 physique (inversion du faisceau appliquee
                                 ensuite, comme dans le pipeline),
                 'extraction' -> elle suit deja l'ordre d'extraction.
    theta_sign : +1 ou -1 — inverse le signe de theta si la convention du
                 fichier est opposee a celle des trace attendues.
    derive_xyz : si le fichier n'a pas de colonnes x, y, z, les calculer en
                 convention spherique standard
                 x = sin θ cos φ, y = sin θ sin φ, z = cos θ.

    Retourne (entree_de_registre, rapport) ou l'entree a la forme
    {"fibres": {...}, "xyz": {...}, "numbering": ..., "source": ...}.
    """
    if n_fibers is None:
        n_fibers = sf.N_FIBERS
    info = sniff_flat_table(path, sheet)
    cols = dict(info["guess"])
    cols.update({k: v for k, v in (col_map or {}).items() if v})
    report = []
    if "fiber" not in cols:
        raise ValueError("Colonne du numero de fibre introuvable — indiquez-la "
                         "manuellement (en-tetes lus : "
                         + ", ".join(info["headers"]) + ").")
    has_ang = "phi" in cols and "theta" in cols
    has_xyz = all(k in cols for k in ("x", "y", "z"))
    if not has_ang and not has_xyz:
        raise ValueError("Ni (phi, theta) ni (x, y, z) n'ont ete trouves — "
                         "indiquez les colonnes manuellement.")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[info["sheet"]]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    head = [str(c).strip() if c is not None else ""
            for c in rows[info["header_row"] - 1]]
    idx = {}
    for field, header in cols.items():
        if header in head:
            idx[field] = head.index(header)
        else:
            report.append(f"Colonne '{header}' ({field}) absente de la feuille "
                          f"— champ ignore.")
    fibres, xyz, seen = {}, {}, set()
    n_bad = 0
    for row in rows[info["header_row"]:]:
        if row is None or all(c is None for c in row):
            continue
        try:
            fnum = int(float(row[idx["fiber"]]))
        except (TypeError, ValueError, IndexError):
            continue
        if not (1 <= fnum <= n_fibers):
            report.append(f"Fibre {fnum} hors de la plage 1–{n_fibers} — ligne "
                          f"ignoree.")
            continue
        if fnum in seen:
            report.append(f"Fibre {fnum} en double — seule la premiere "
                          f"occurrence est gardee.")
            continue
        seen.add(fnum)
        key = fnum - 1                       # 1-based fichier -> 0-based table
        phi = theta = None
        if has_ang:
            try:
                phi = float(row[idx["phi"]]) * float(phi_sign)
                theta = float(row[idx["theta"]]) * float(theta_sign)
            except (TypeError, ValueError, IndexError):
                phi = theta = None
        if phi is None or theta is None:
            n_bad += 1
        else:
            fibres[key] = (phi, theta)
        if has_xyz:
            try:
                xyz[key] = (float(row[idx["x"]]), float(row[idx["y"]]),
                            float(row[idx["z"]]))
            except (TypeError, ValueError, IndexError):
                pass
        elif derive_xyz and phi is not None:
            th, ph = np.radians(theta), np.radians(phi)
            xyz[key] = (float(np.sin(th) * np.cos(ph)),
                        float(np.sin(th) * np.sin(ph)),
                        float(np.cos(th)))
    if n_bad:
        report.append(f"{n_bad} fibres sans angle exploitable (NaN sur les "
                      f"cartes).")
    missing = [i for i in range(1, n_fibers + 1) if i not in seen]
    if missing:
        report.append(f"{len(missing)} fibres absentes du fichier "
                      f"({', '.join(map(str, missing[:10]))}"
                      + ("…" if len(missing) > 10 else "") + ").")
    if has_xyz:
        report.append("Coordonnees x, y, z lues directement dans le fichier.")
    elif derive_xyz:
        report.append("Pas de colonnes x, y, z : positions recalculees depuis "
                      "(phi, theta) en convention spherique standard.")
    if quadrants:
        found = sorted(set(quadrants.values()))
        rule = [q for q in found if q.startswith(str(negate or ""))] \
            if negate else []
        report.append(("info",
                       "Quadrants detected: " + ", ".join(found)
                       + ". Initial sign convention: "
                       + (f"θ negated for {', '.join(rule)}" if rule
                          else "no inversion")
                       + " — changeable in the 3D view (page 7)."))
    else:
        rule = []
    ent = {"fibres": fibres, "xyz": xyz, "numbering": numbering,
           "source": Path(path).name, "quadrant": quadrants,
           "sign_rules": normalise_sign_rules({"theta": rule, "phi": []})}
    return ent, report


def flat_table_summary(ent: dict) -> dict:
    """Quelques chiffres pour l'affichage de controle."""
    e = entry(ent)
    fib = e["fibres"]
    if not fib:
        return {"n": 0}
    phis = np.array([v[0] for v in fib.values()], float)
    thetas = np.array([v[1] for v in fib.values()], float)
    return {"n": len(fib), "n_xyz": len(e["xyz"] or {}),
            "phi_min": float(np.nanmin(phis)), "phi_max": float(np.nanmax(phis)),
            "theta_min": float(np.nanmin(thetas)),
            "theta_max": float(np.nanmax(thetas)),
            "numbering": e["numbering"], "source": e.get("source", "")}


# =============================================================================
# Plan de fibres de campagne — quadrant / bras / port UNIQUEMENT
# =============================================================================
# Constat de terrain : d'une campagne a l'autre, le classeur decrivant les
# fibres change de nom, de colonnes et d'ordre, et les angles (phi, theta) ou
# les coordonnees (x, y, z) qu'il contient parfois se sont averes FAUX
# (convention theta -> 180-theta observee sur fiber_config_Fauvel_b.xlsx, ce
# qui retourne la sphere en z).
#
# La seule chose stable est le fichier de reference des angles
# (…EstimateStructureCoordinates.xlsx). On ne lit donc du fichier de campagne
# que les trois informations qui ne peuvent pas mentir :
#
#       numero de fibre | quadrant (alpha/delta) | bras (lettre) | port (n°)
#
# puis on va chercher phi(bras) et theta(port) dans le fichier de reference et
# on recalcule x, y, z. Les colonnes phi/theta/x/y/z eventuellement presentes
# dans le fichier de campagne ne sont JAMAIS utilisees : elles sont seulement
# relues pour signaler l'ecart a l'utilisateur.
#
# La detection des colonnes se fait d'abord par l'intitule, puis — si les
# intitules sont inconnus — PAR LE CONTENU (un quadrant contient "alpha" ou
# "delta", un bras est une lettre seule, un port un petit entier, une fibre un
# entier unique de 1 a N). Une campagne future peut donc renommer ou deplacer
# ses colonnes sans qu'il y ait une ligne de code a changer.

MAP_FIELDS = ["fiber", "quadrant", "arm", "port"]

MAP_ALIASES = {
    "fiber": ["fiber", "fiber #", "fiber number", "fiber id", "fibre",
              "fibre #", "fibre n", "channel", "chan", "num", "no", "n",
              "id", "index"],
    "quadrant": ["quadrant", "cadran", "config type", "config", "configuration",
                 "type", "structure", "hemisphere", "cone", "side", "group",
                 "groupe", "family"],
    "arm": ["arm", "bras", "arm letter", "letter", "lettre", "branch",
            "branche", "leg"],
    "port": ["port", "port #", "port number", "port n", "position", "pos",
             "ring", "anneau", "slot"],
}

_QUAD_RE = re.compile(r"(alpha|delta)", re.IGNORECASE)
_SINGLE_LETTER_RE = re.compile(r"^[A-Za-z]$")
_TRAILING_LETTER_RE = re.compile(r"(?:^|[\s_\-])([A-Za-z])\s*$")


def _cell_str(v) -> str:
    return "" if v is None else str(v).strip()


def _as_int(v):
    """Entier si la cellule en est un (12 ou 12.0), sinon None."""
    if isinstance(v, bool):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(f) or abs(f - round(f)) > 1e-9:
        return None
    return int(round(f))


def _as_float(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if np.isfinite(f) else None


def quadrant_of(text):
    """'Delta I', 'delta', 'DELTA_ARM' -> 'delta'. None si indecidable."""
    m = _QUAD_RE.search(_cell_str(text))
    return m.group(1).lower() if m else None


def arm_of(text):
    """Lettre de bras d'une cellule : 'I' -> 'I', 'Delta I' -> 'I'.
    'Delta' seul ne donne PAS 'a' (la lettre doit etre un mot a elle seule)."""
    t = _cell_str(text)
    if _SINGLE_LETTER_RE.match(t):
        return t.upper()
    m = _TRAILING_LETTER_RE.search(t)
    return m.group(1).upper() if m else None


def _find_header_row(rows, max_scan=15):
    """Ligne d'en-tete = premiere ligne d'au moins 3 cellules suivie d'une
    ligne de donnees d'au moins 3 cellules contenant au moins un nombre."""
    for i in range(min(max_scan, max(0, len(rows) - 1))):
        head = [c for c in rows[i] if _cell_str(c)]
        nxt = rows[i + 1]
        nxt_filled = [c for c in nxt if _cell_str(c)]
        if len(head) >= 3 and len(nxt_filled) >= 3 \
                and any(_as_float(c) is not None for c in nxt):
            return i
    return 0


def _column_stats(rows, ncol):
    out = []
    for j in range(ncol):
        vals = [r[j] if j < len(r) else None for r in rows]
        vals = [v for v in vals if _cell_str(v)]
        n = max(len(vals), 1)
        ints = [_as_int(v) for v in vals]
        ints = [v for v in ints if v is not None]
        out.append({
            "j": j,
            "n": len(vals),
            "f_int": len(ints) / n,
            "f_quad": sum(1 for v in vals if quadrant_of(v)) / n,
            "f_arm": sum(1 for v in vals
                         if _SINGLE_LETTER_RE.match(_cell_str(v))) / n,
            "f_arm_tail": sum(1 for v in vals if arm_of(v)) / n,
            "ints": ints,
        })
    return out


def _guess_map_columns(headers, stats):
    """{champ: index de colonne}. Intitules d'abord, contenu ensuite."""
    guess, used = {}, set()

    # 1) par intitule
    norm = {j: _norm_header(h) for j, h in enumerate(headers)}
    for field in MAP_FIELDS:
        for alias in sorted(MAP_ALIASES[field], key=len, reverse=True):
            for j, t in norm.items():
                if j in used or not t:
                    continue
                if t == alias or t.startswith(alias + " ") \
                        or t.endswith(" " + alias):
                    guess[field] = j
                    used.add(j)
                    break
            if field in guess:
                break

    # 2) par contenu — ce qui rend le lecteur independant des intitules
    if "quadrant" not in guess:
        cands = [s for s in stats if s["f_quad"] >= 0.8 and s["j"] not in used]
        if cands:
            j = max(cands, key=lambda s: s["f_quad"])["j"]
            guess["quadrant"] = j
            used.add(j)
    if "fiber" not in guess:
        for s in sorted(stats, key=lambda s: s["j"]):
            if s["j"] in used or s["f_int"] < 0.95 or not s["ints"]:
                continue
            iv = s["ints"]
            if len(set(iv)) == len(iv) and min(iv) == 1 and max(iv) == len(iv):
                guess["fiber"] = s["j"]
                used.add(s["j"])
                break
    if "port" not in guess:
        for s in sorted(stats, key=lambda s: s["j"]):
            if s["j"] in used or s["f_int"] < 0.95 or not s["ints"]:
                continue
            if min(s["ints"]) >= 1 and max(s["ints"]) <= 60:
                guess["port"] = s["j"]
                used.add(s["j"])
                break
    if "arm" not in guess:
        cands = [s for s in stats if s["f_arm"] >= 0.8 and s["j"] not in used]
        if cands:
            j = max(cands, key=lambda s: s["f_arm"])["j"]
            guess["arm"] = j
            used.add(j)

    # 3) colonne combinee « Delta I » : le bras se lit dans le quadrant
    if "arm" not in guess and "quadrant" in guess:
        s = stats[guess["quadrant"]]
        if s["f_arm_tail"] >= 0.8:
            guess["arm"] = guess["quadrant"]
    return guess


def sniff_fiber_map(path, sheet=None) -> dict:
    """Inspecte un classeur de campagne sans rien decider en silence.

    Retourne {sheets, sheet, header_row (1-based), headers, guess
    {champ: intitule}, guess_idx {champ: index}, n_rows, preview}.
    La feuille est choisie automatiquement : celle qui donne le plus de
    champs reconnus (une feuille « Summary » ne gagne pas contre la table
    des fibres).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = list(wb.sheetnames)
    names = [sheet] if sheet in sheets else sheets
    best = None
    for name in names:
        ws = wb[name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(_cell_str(c) for c in r)]
        if len(rows) < 2:
            continue
        h = _find_header_row(rows)
        headers = [_cell_str(c) for c in rows[h]]
        data = rows[h + 1:]
        ncol = max([len(r) for r in data] + [len(headers)])
        stats = _column_stats(data, ncol)
        g = _guess_map_columns(headers, stats)
        score = (len([f for f in MAP_FIELDS if f in g]), len(data))
        if best is None or score > best["score"]:
            best = {"score": score, "sheet": name, "header_row": h + 1,
                    "headers": headers, "guess_idx": g, "n_rows": len(data),
                    "preview": data[:5], "rows": data}
    wb.close()
    if best is None:
        raise ValueError("No usable sheet in this workbook.")
    heads = best["headers"]
    best["guess"] = {f: (heads[j] if j < len(heads) and heads[j]
                         else f"col {j + 1}")
                     for f, j in best["guess_idx"].items()}
    best["sheets"] = sheets
    best.pop("score", None)
    return best


def _own_angle_columns(headers):
    """Index des colonnes phi/theta/x/y/z du fichier de campagne — lues
    UNIQUEMENT pour signaler l'ecart, jamais pour calculer."""
    idx = {}
    norm = {j: _norm_header(h) for j, h in enumerate(headers)}
    for field in ("phi", "theta", "x", "y", "z"):
        for alias in sorted(FLAT_ALIASES[field], key=len, reverse=True):
            for j, t in norm.items():
                if j in idx.values() or not t:
                    continue
                if t == alias or t.startswith(alias + " "):
                    idx[field] = j
                    break
            if field in idx:
                break
    return idx


def build_config_from_map(structure: dict, path, sheet=None, col_map=None,
                          negate=NEGATE_QUADRANT, numbering="extraction",
                          n_fibers=None):
    """Construit une configuration angulaire a partir d'un plan de fibres.

    `structure` : tables renvoyees par load_structure_tables() — la SEULE
    source d'angles. `col_map` : {champ: index de colonne} pour forcer la
    correspondance si la detection automatique se trompe.

    Retourne (entree_de_registre, rapport) ou le rapport est une liste de
    (niveau, message) avec niveau dans {'ok', 'info', 'warn'}.
    """
    if n_fibers is None:
        n_fibers = sf.N_FIBERS
    info = sniff_fiber_map(path, sheet)
    idx = dict(info["guess_idx"])
    idx.update({k: int(v) for k, v in (col_map or {}).items()
                if v is not None and str(v) != ""})
    report = []
    missing_fields = [f for f in MAP_FIELDS if f not in idx]
    if missing_fields:
        raise ValueError(
            "Columns not found in '" + Path(path).name + "': "
            + ", ".join(missing_fields)
            + ". Headers read: " + ", ".join(h for h in info["headers"] if h)
            + ". Set them under 'Manual override'.")

    own = _own_angle_columns(info["headers"])
    fibres, xyz, quadrants = {}, {}, {}
    arms_used, ports_used = {}, set()
    seen, extrapolated, problems = set(), [], []
    dphi_max = dtheta_max = 0.0
    dxyz_max = 0.0

    for row in info["rows"]:
        def cell(field):
            j = idx[field]
            return row[j] if j < len(row) else None

        fnum = _as_int(cell("fiber"))
        if fnum is None or not (1 <= fnum <= n_fibers):
            continue
        if fnum in seen:
            problems.append(f"fiber {fnum} appears twice — first row kept")
            continue
        seen.add(fnum)

        quad = quadrant_of(cell("quadrant"))
        arm = arm_of(cell("arm"))
        port = _as_int(cell("port"))
        if quad is None:
            problems.append(f"fiber {fnum}: unreadable quadrant "
                            f"({cell('quadrant')!r})")
            continue
        if arm is None:
            problems.append(f"fiber {fnum}: unreadable arm ({cell('arm')!r})")
            continue
        if port is None:
            problems.append(f"fiber {fnum}: unreadable port ({cell('port')!r})")
            continue

        phi = structure["phi_" + quad].get(arm)
        th, extrap = _theta_lookup(structure["th_" + quad], port)
        if phi is None:
            problems.append(f"fiber {fnum}: arm '{arm}' missing from the "
                            f"phi table ({quad}) of the reference file")
            continue
        if th is None:
            problems.append(f"fiber {fnum}: port {port} missing from the "
                            f"theta table ({quad}), which is not "
                            f"extrapolable")
            continue
        if extrap:
            extrapolated.append((fnum, port, th))

        # Les angles sont stockes BRUTS. Le changement de signe eventuel est
        # une convention de montage, choisie a l'affichage (page 7) parmi les
        # quadrants reellement presents — voir _apply_sign.
        arms_used.setdefault((quad, arm), phi)
        ports_used.add(port)
        quadrants[fnum - 1] = quad
        fibres[fnum - 1] = (float(phi), float(th))
        cx, cy, cz = (float(c) for c in
                      xyz_from_angles(phi, th))
        xyz[fnum - 1] = (cx, cy, cz)

        # Comparaison (informative) avec ce que le fichier pretend
        if "phi" in own and "theta" in own:
            fp = _as_float(row[own["phi"]] if own["phi"] < len(row) else None)
            ft = _as_float(row[own["theta"]] if own["theta"] < len(row) else None)
            if fp is not None:
                dphi_max = max(dphi_max, abs(((fp - phi + 180) % 360) - 180))
            if ft is not None:
                dtheta_max = max(dtheta_max, abs(ft - th))
        if all(k in own for k in ("x", "y", "z")):
            vals = [_as_float(row[own[k]]) if own[k] < len(row) else None
                    for k in ("x", "y", "z")]
            if all(v is not None for v in vals):
                dxyz_max = max(dxyz_max, float(np.max(np.abs(
                    np.array(vals) - np.array([cx, cy, cz])))))

    # ── Rapport ────────────────────────────────────────────────────────────
    report.append(("ok", f"Sheet '{info['sheet']}', header on row "
                         f"{info['header_row']}, {info['n_rows']} rows read."))
    report.append(("ok", "Columns used — "
                   + ", ".join(f"{f}: '{info['guess'].get(f, '?')}'"
                               if idx[f] == info["guess_idx"].get(f)
                               else f"{f}: column {idx[f] + 1} (forced)"
                               for f in MAP_FIELDS)))
    report.append(("ok" if len(fibres) == n_fibers else "warn",
                   f"{len(fibres)}/{n_fibers} fibers resolved through the "
                   f"angle reference file."))
    if arms_used:
        desc = ", ".join(f"{q[:1].upper()}{q[1:]} {a} → φ={v:g}°"
                         for (q, a), v in sorted(arms_used.items()))
        report.append(("info", f"Arms recognised: {desc}."))
    if ports_used:
        report.append(("info", f"Ports used: {min(ports_used)}–"
                               f"{max(ports_used)}."))
    if extrapolated:
        ex = ", ".join(f"fiber {f} (port {p} → {t:+.2f}°)"
                       for f, p, t in extrapolated[:6])
        report.append(("warn", f"{len(extrapolated)} ports beyond the theta "
                               f"table, extrapolated with the constant step: "
                               f"{ex}"
                       + ("…" if len(extrapolated) > 6 else "")))
    missing = [i for i in range(1, n_fibers + 1) if i not in seen]
    if missing:
        report.append(("warn", f"{len(missing)} fibers missing from the file "
                               f"({', '.join(map(str, missing[:10]))}"
                       + ("…" if len(missing) > 10 else "")
                       + ") — NaN on the maps."))
    for p in problems[:10]:
        report.append(("warn", p))
    if len(problems) > 10:
        report.append(("warn", f"… and {len(problems) - 10} further "
                               f"problematic rows."))
    if "phi" in own or "theta" in own:
        if max(dphi_max, dtheta_max) > 1e-6:
            report.append(("warn",
                           f"The campaign file carries its own angles; they "
                           f"depart from the reference by {dphi_max:.2f}° in "
                           f"φ and {dtheta_max:.2f}° in θ. They were IGNORED "
                           f"(the reference file is authoritative)."))
        else:
            report.append(("ok", "The angles written in the campaign file "
                                 "match the reference."))
    if all(k in own for k in ("x", "y", "z")) and dxyz_max > 1e-6:
        report.append(("warn", f"The x, y, z coordinates in the file differ "
                               f"by {dxyz_max:.3f} (unit radius) from the "
                               f"recomputed ones — they were IGNORED."))
    report.append(("info", f"Fiber numbering: '{numbering}' — the file's "
                          f"fiber numbers are read as "
                          + ("detector/extraction order (no beam inversion)."
                             if str(numbering).startswith("extract")
                             else "physical fiber numbers (beam inversion "
                                  "applied).")))
    report.append(("info", "x, y, z recomputed from (φ, θ) about the BEAM "
                           "axis: x = cosθ, y = sinθ·cosφ, z = sinθ·sinφ. "
                           "θ is the scattering angle from backscatter (+x); "
                           "φ turns around the beam from +y."))

    if quadrants:
        found = sorted(set(quadrants.values()))
        rule = [q for q in found if q.startswith(str(negate or ""))] \
            if negate else []
        report.append(("info",
                       "Quadrants detected: " + ", ".join(found)
                       + ". Initial sign convention: "
                       + (f"θ negated for {', '.join(rule)}" if rule
                          else "no inversion")
                       + " — changeable in the 3D view (page 7)."))
    else:
        rule = []
    ent = {"fibres": fibres, "xyz": xyz, "numbering": numbering,
           "source": Path(path).name, "quadrant": quadrants,
           "sign_rules": normalise_sign_rules({"theta": rule, "phi": []})}
    return ent, report
